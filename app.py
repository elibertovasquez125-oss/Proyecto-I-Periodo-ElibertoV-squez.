# Crea la aplicación web y maneja las páginas y formularios en el navegador
from flask import Flask, render_template, request, redirect, url_for

# Revisa si los archivos y carpetas ya existen en la computadora
import os

# Crea archivos de Excel nuevos y abre el que ya existe
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

ARCHIVO_EXCEL = "contactos.xlsx"

# Crear archivo Excel si no existe
if not os.path.exists(ARCHIVO_EXCEL):
    
    # Crea un libro de Excel nuevo en blanco
    libro = Workbook()
    # Selecciona la primera hoja de ese libro para empezar a trabajar en ella
    hoja = libro.active

    # Le cambia el nombre a la hoja para que se llame "Contactos"
    hoja.title = "Contactos"
    # escribimos los titulos de la primera fila
    hoja.append(["Nombre","Apellido","Telefono","Correo","Direccion","Categoria","Favorito"])

    # guardamos el archivo
    libro.save(ARCHIVO_EXCEL)


# Ruta principal
@app.route('/')
def home():
    # Envía automáticamente al usuario a la página de inicio de sesión
    return redirect(url_for('login'))


# Login
@app.route('/login', methods=['GET', 'POST'])
def login():
    
    # Verifica si el usuario presionó el botón para enviar el formulario
    if request.method == 'POST':

        # Recoge el texto que escribio en usuario y contraseña
        usuario = request.form['usuario']
        contraseña = request.form['contraseña']

        # comprobamos que se los datos que ya habiamos predefinido
        if usuario == "admin" and contraseña == "1234":
            return redirect(url_for('inicio'))

        # si no coinciden
        else:
            # vuelve a cargar la pagina y dara un mensaje
            return render_template(
                'login.html',
                error='Usuario o contraseña incorrecta'
            )

    return render_template('login.html')


# Página principal que se define la ruta
@app.route('/casa')
def inicio():

    # abre el archivo de excel ya existente
    libro = load_workbook(ARCHIVO_EXCEL)
    # se selecciona la hoja ya activa
    hoja = libro.active

    # crea una lista vacia para guardar los contactos
    contactos = []

    # Saltar encabezados para no leer los titulos y leer solo datos de contactos
    for fila in hoja.iter_rows(min_row=2, values_only=True):

        # Transforma los datos de cada celda en un diccionario ordenado y lo añade a la lista
        contactos.append({
        "nombre": fila[0],
        "apellido": fila[1],
        "telefono": fila[2],
        "correo": fila[3],
        "direccion": fila[4],
        "categoria": fila[5],
        "favorito": fila[6]
        })
        

    # se carga la página y se la pasa al excel para que despues se vea en la pagina
    return render_template('index.html', contactos=contactos)


# Buscar contacto
@app.route('/buscar', methods=['GET'])
def buscar():

    # obtenemos el texto del buscador si esta mayuscula lo pasa minusculas
    nombre_buscado = request.args.get('nombre', '').lower()

    # el archivo de excel existente y selecciona la hoja donde donde se guardo
    libro = load_workbook(ARCHIVO_EXCEL)
    hoja = libro.active

    # crea una lista que guarda los contactos que cumplan con la busqueda
    resultados = []

    # recorre el excel fila por fila a partir de la fila 2
    for fila in hoja.iter_rows(min_row=2, values_only=True):

        # verifica que el nombre que se busca esta dentro de nombre de la fila
        if nombre_buscado in str(fila[0]).lower():

            # si coincide guarda los datos de este contacto organizodo dentro de la lista resusltado
            resultados.append({
                "nombre": fila[0],
                "apellido": fila[1],
                "telefono": fila[2],
                "correo": fila[3],
                "direccion": fila[4],
                "categoria": fila[5],
                "favorito": fila[6]
            })
    return render_template(
        'index.html',
        contactos=resultados
    )

# defini la ruta ordenar
@app.route('/ordenar')
def ordenar():

    libro = load_workbook(ARCHIVO_EXCEL)
    hoja = libro.active

    # crea una lista para almacena por un tiempo los contactos
    contactos = []

    for fila in hoja.iter_rows(min_row=2, values_only=True):

        # agrega la información de cada fila de forma ordenada
        contactos.append({
            "nombre": fila[0],
            "apellido": fila[1],
            "telefono": fila[2],
            "correo": fila[3],
            "direccion": fila[4],
            "categoria": fila[5],
            "favorito": fila[6]
        })
    # ordena la lista alfabeticamente (A-Z)
    contactos.sort(
        key=lambda contacto: contacto["nombre"].lower()
    )

    return render_template('index.html', contactos=contactos)

# Agregar contacto
@app.route('/agregar', methods=['POST'])
def agregar():

    # Recibe y guarda la información enviada desde cada cuadro de texto del formulario
    nombre = request.form['nombre']
    apellido = request.form['apellido']
    telefono = request.form['telefono']
    correo = request.form['correo']
    direccion = request.form['direccion']
    categoria = request.form['categoria']
    favorito = request.form['favorito']

    libro = load_workbook(ARCHIVO_EXCEL)
    hoja = libro.active

    # añade los datos obtenidos como una nueva fila al final de la lista de excel
    hoja.append([nombre, apellido, telefono, correo,direccion, categoria, favorito ])

    libro.save(ARCHIVO_EXCEL) #guardamos los cambios

    return redirect(url_for('inicio')) #se recar la página y redirige a la pagian principal

# mas información del contacto
@app.route('/mas_informacion', methods=['POST'])
def mas_informacion():

    # recibe telefono de contacto seleccionado desde formulario
    telefono = request.form['telefono']

    libro = load_workbook(ARCHIVO_EXCEL)
    hoja = libro.active

    contacto = None #variable vacia que guarda la informacion de contacto una vez que se encontro

    for fila in hoja.iter_rows(min_row=2, values_only=True):

        # revisa que el telefono coincida con el telefono recibido
        if str(fila[2]) == telefono:

            contacto = {
                "nombre": fila[0],
                "apellido": fila[1],
                "telefono": fila[2],
                "correo": fila[3],
                "direccion": fila[4],
                "categoria": fila[5],
                "favorito": fila[6]
            }
            # detiene el ciclo por que ya se encontro el contacto
            break

    # carga la pagina
    return render_template(
        'mas_informacion.html',
        contacto=contacto
    )



# eliminar contacto ya existente
@app.route('/eliminar', methods=['POST'])
def eliminar():
    # Recibe el nombre y el teléfono ingresados en el formulario de eliminación
    nombre = request.form['nombre']
    telefono = request.form['telefono']

    libro = load_workbook(ARCHIVO_EXCEL)
    hoja = libro.active

    for fila in hoja.iter_rows(min_row=2):

        # comprueba si le dato coincide con el nombre ingresado
        if fila[0].value == nombre:

            hoja.delete_rows(fila[0].row)
            break
        # sino comprueba con el numero de telefono ingresado
        elif fila[2].value == telefono:
            hoja.delete_rows(fila[0].row)
            # detiene el cilo cuando se borre el contacto
            break
    # guardamos los cambios
    libro.save(ARCHIVO_EXCEL)

    return redirect(url_for('inicio')) #recargamos la página

# editar contacto
@app.route('/editar', methods=['POST'])
def editar():

    # Recibe el teléfono (llave para buscar) y los nuevos datos ingresados en el formulario
    telefono = request.form['telefono']
    nuevo_nombre = request.form['nuevo_nombre']
    nuevo_apellido = request.form['nuevo_apellido']
    nuevo_correo = request.form['nuevo_correo']
    nueva_direccion = request.form['nueva_direccion']

    libro = load_workbook(ARCHIVO_EXCEL)
    hoja = libro.active

    for fila in hoja.iter_rows(min_row=2):

        # comprueba que el dato de la fila sea igual al numero ingresado
        if str(fila[2].value) == telefono:
            
            # se reemplaza los datos anteriores por los nuevos ingresados
            fila[0].value = nuevo_nombre
            fila[1].value = nuevo_apellido
            fila[3].value = nuevo_correo
            fila[4].value = nueva_direccion
            
            # finaliza el ciclo cuando se haga todo
            break
    # realizamos el guardado de los cambios
    libro.save(ARCHIVO_EXCEL)

    return redirect(url_for('inicio')) #recargamos la página


# generar un reporte de contactos registrados
@app.route('/reporte')
def reporte():

    libro = load_workbook(ARCHIVO_EXCEL)
    hoja = libro.active

    # realizamos un contados para saber cuantos contactos y favoritos hay
    total_contactos = 0
    total_favoritos = 0

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        
        # suma q por cada fila que hay un contacto
        total_contactos += 1

        # comprobamos la fila 6 que dato hay y sumara 1 si es "Si"
        if str(fila[6]).lower() == "si":
            total_favoritos += 1

    # carga la página dsepus de realizar todo
    return render_template(
        'reporte.html',
        total_contactos=total_contactos,
        total_favoritos=total_favoritos
    )

# comprueba si este archivo de Python se está ejecutando directamente
if __name__ == '__main__':
    # enciende el servidor web de flask y activa el modod de prubas 
    app.run(debug=True)